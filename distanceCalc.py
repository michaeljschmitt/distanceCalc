import openpyxl
from geopy.distance import great_circle

option = "closest" # can choose between "closest" distances to set A coordinates, or "all" distances between the two sets of coordinates

wb = openpyxl.load_workbook("Data.xlsx")
ws = wb.active

wbOutput = openpyxl.Workbook()
wsOutput = wbOutput.active

setACoordinates = []
setBCoordinates = []
pointsByDistance = []

setATitle = ws["A1"].value
setBTitle = ws["D1"].value

for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
	
	setAID = row[0].value
	latitude = row[1].value
	longitude = row[2].value
	coord = (latitude, longitude)
	setACoordinates.append([setAID, coord])

for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
	
	setBID = row[0].value
	latitude = row[1].value
	longitude = row[2].value
	coord = (latitude, longitude)
	setBCoordinates.append([setBID, coord])

for pointA in setACoordinates:
	for pointB in setBCoordinates:
		distance = great_circle(pointA[1],pointB[1]).miles
		pointsByDistance.append([pointA[0], pointB[0], distance])

shortestPointByDistance = {}
for pointByDistance in pointsByDistance:
	setAPoint = pointByDistance[0]
	setBPoint = pointByDistance[1]
	distance = pointByDistance[2]

	if distance != 0:
		if setAPoint == None:
			continue
		elif setAPoint not in shortestPointByDistance.keys():
			shortestPointByDistance[setAPoint] = [setBPoint, distance]
		else:
			currentShortestDistance = shortestPointByDistance[setAPoint][1]
			if distance < currentShortestDistance:
				shortestPointByDistance[setAPoint] = [setBPoint, distance]

# convert dictionary to list
closestPointsByDistance = []
for key, value in shortestPointByDistance.items():
	setAPoint = key
	setBPoint = value[0]
	distance = value[1]

	closestPointsByDistance.append([setAPoint,setBPoint,distance])
	
wsOutput.append([setATitle, setBTitle, 'Distance (mi)'])

# for closest distances (to point A coordinates)
if option == 'closest':
	for pointByDistance in closestPointsByDistance:
		setAPoint = pointByDistance[0]
		setBPoint = pointByDistance[1]
		distance = pointByDistance[2]

		wsOutput.append([setAPoint, setBPoint, distance])

# for all distances
if option == 'all':
	for pointByDistance in pointsByDistance:
		setAPoint = pointByDistance[0]
		setBPoint = pointByDistance[1]
		distance = pointByDistance[2]
		
		wsOutput.append([setAPoint, setBPoint, distance])

wbOutput.save('./Output.xlsx')
